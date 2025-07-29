"""TK INTER IS INTEGRATED WHICH WILL EXTRACT THE FORMULA OF ANY TABLE OF EXCEL AND DELIVER IT TO THE TABLE AREA CALCULATION FORMULA TABLE"""
"""Extract Excel formulas and store them in SQL Server - Sperene Tools"""

import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import pyodbc
import traceback

# --- CONFIG ---
DB_CONFIG = {
    'server': 'DESKTOP-PLHJ2M7',
    'database': 'test',
    'trusted_connection': 'yes'
}

# --- UTILITIES ---
def extract_formulas(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    sheet = workbook.active
    formulas = []
    headers = [cell.value for cell in next(sheet.iter_rows())]

    for row in sheet.iter_rows(min_row=2):
        for col_index, cell in enumerate(row):
            if cell.data_type == 'f':
                formula_info = {
                    'Formula': cell.value,
                    'Header': headers[col_index],
                    'CellName': cell.coordinate,
                    'InvolvedColumns': replace_with_headers(cell.value, headers)
                }
                formulas.append(formula_info)

    workbook.close()
    return formulas

def replace_with_headers(formula, headers):
    column_refs = re.findall(r'([A-Za-z]+)(\d+)', formula)
    for col_ref, row in column_refs:
        col_index = openpyxl.utils.cell.column_index_from_string(col_ref) - 1
        if col_index < len(headers) and headers[col_index]:
            formula = formula.replace(f"{col_ref}{row}", headers[col_index])
    return formula

def create_tracking_table_if_not_exists(cursor):
    cursor.execute('''
        IF NOT EXISTS (SELECT * FROM sys.tables WHERE name='areaCalculationSheet_Formulas')
        BEGIN
            CREATE TABLE areaCalculationSheet_Formulas (
                Formula NVARCHAR(MAX),
                Column_Header NVARCHAR(255),
                CellName NVARCHAR(255),
                Involved_Columns NVARCHAR(MAX),
                Table_Name NVARCHAR(255)
            );
        END
    ''')

def update_or_insert(cursor, table_name, formulas):
    for item in formulas:
        cursor.execute('''
            IF EXISTS (SELECT * FROM areaCalculationSheet_Formulas WHERE CellName = ? AND Table_Name = ?)
            BEGIN
                UPDATE areaCalculationSheet_Formulas
                SET Formula = ?, Column_Header = ?, Involved_Columns = ?
                WHERE CellName = ? AND Table_Name = ?
            END
            ELSE
            BEGIN
                INSERT INTO areaCalculationSheet_Formulas (Formula, Column_Header, CellName, Involved_Columns, Table_Name)
                VALUES (?, ?, ?, ?, ?)
            END
        ''', (
            item['CellName'], table_name,
            item['Formula'], item['Header'], item['InvolvedColumns'],
            item['CellName'], table_name,
            item['Formula'], item['Header'], item['CellName'], item['InvolvedColumns'], table_name
        ))

def store_to_db(formulas, table_name):
    conn_str = f"DRIVER={{SQL Server}};SERVER={DB_CONFIG['server']};DATABASE={DB_CONFIG['database']};Trusted_Connection={DB_CONFIG['trusted_connection']}"
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()
    create_tracking_table_if_not_exists(cursor)
    update_or_insert(cursor, table_name, formulas)
    conn.commit()
    conn.close()

# --- GUI LOGIC ---
def process_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
    if not file_path:
        status_var.set("⚠️ No file selected.")
        selected_file_var.set("No file selected.")
        return

    file_name = os.path.splitext(os.path.basename(file_path))[0]
    selected_file_var.set(f"📄 Selected File: {os.path.basename(file_path)}")

    try:
        formulas = extract_formulas(file_path)
        if formulas:
            store_to_db(formulas, file_name)
            msg = f"✅ '{file_name}' processed.\n📊 {len(formulas)} formula(s) stored to DB."
            status_var.set(msg)
            messagebox.showinfo("Success", msg)
        else:
            status_var.set("ℹ️ No formulas found in the Excel file.")
            messagebox.showinfo("Info", "No formulas were found in the selected file.")
    except Exception as e:
        error_msg = f"❌ Error: {e}"
        status_var.set(error_msg)
        traceback.print_exc()
        messagebox.showerror("Error", f"An error occurred:\n{e}")

# --- GUI DESIGN ---
def launch_gui():
    global status_var, selected_file_var

    root = tk.Tk()
    root.title("Excel Formula Extractor - Sperene Tools")
    root.geometry("750x580")
    root.config(bg="#f0f8ff")

    # Header
    tk.Label(root, text="🧮 Excel Formula Extractor", font=("Helvetica", 22, "bold"),
             bg="#4682B4", fg="white", pady=14).pack(fill=tk.X)

    # Subtext
    tk.Label(
        root,
        text="🔎 Select an Excel file to extract formulas used in its cells.\n"
             "📤 The extracted data will be stored in SQL Server (table: areaCalculationSheet_Formulas).",
        font=("Arial", 11), bg="#f0f8ff", fg="#333", justify="center"
    ).pack(pady=10)

    # Browse Button
    ttk.Button(root, text="📂 Browse Excel File", command=process_excel_file).pack(pady=12, ipadx=14, ipady=7)

    # Selected File Label
    selected_file_var = tk.StringVar()
    selected_file_var.set("No file selected.")
    tk.Label(root, textvariable=selected_file_var, bg="#f0f8ff", fg="#444", font=("Arial", 10, "italic")).pack()

    # Info Frame
    info_frame = tk.Frame(root, bg="#f0f8ff", padx=30, pady=10)
    info_frame.pack(pady=5)

    # DB Info
    db_label = f"🗄️ Connected to: {DB_CONFIG['database']} @ {DB_CONFIG['server']}"
    tk.Label(info_frame, text=db_label, bg="#f0f8ff", fg="#005f5f", font=("Arial", 10, "italic")).pack(anchor="w", pady=2)

    # Features
    tk.Label(info_frame, text="✨ What You Can Do:", font=("Arial", 10, "bold"),
             bg="#f0f8ff", fg="#333").pack(anchor="w", pady=(10, 2))

    features = [
        "✔️ Extract formulas used in Excel sheets",
        "✔️ See which columns the formulas refer to",
        "✔️ Automatically stores data into SQL Server",
        "✔️ No data lost — auto-update if formula changes"
    ]
    for feat in features:
        tk.Label(info_frame, text=feat, font=("Arial", 10), bg="#f0f8ff", anchor="w").pack(anchor="w")

    # Filler
    tk.Label(root, text="", bg="#f0f8ff").pack(expand=True)

    # Status Bar
    status_var = tk.StringVar()
    tk.Label(root, textvariable=status_var, anchor="w", bg="#d3eaff", font=("Arial", 11)).pack(fill=tk.X, side=tk.BOTTOM, ipady=6)

    # Footer
    tk.Label(root, text="🧠 Powered by Sperene Technologies", bg="#d3eaff", fg="#333",
             font=("Arial", 9, "italic")).pack(side=tk.BOTTOM, fill=tk.X)

    root.mainloop()

# --- SAFELY LAUNCH ---
if __name__ == "__main__":
    try:
        launch_gui()
    except Exception as e:
        traceback.print_exc()
        input(f"\n[ERROR OCCURRED] {e}\nPress Enter to exit...")


