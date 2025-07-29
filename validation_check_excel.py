"""THIS IS FOR TO CHECK IF THE REQUIRED FILE HAS FORMULA USED ON THE REQUIRED PLACES / IN THE PROVIDED EXCEL WHERE HAS BEEN FORMULA USED """
import openpyxl
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import subprocess

# ------------------- Highlight Logic -------------------
def highlight_formula_cells(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        highlight = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        formula_count = 0
        for row in sheet.iter_rows(min_row=2):  # Skip header
            for cell in row:
                if cell.data_type == 'f':
                    cell.fill = highlight
                    formula_count += 1

        wb.save(file_path)

        # Automatically open the file
        try:
            os.startfile(file_path)  # Windows only
        except Exception as e:
            subprocess.Popen(['open', file_path])  # Mac/Linux fallback

        return f"✅ {formula_count} formula cell(s) highlighted successfully!"
    except Exception as e:
        return f"❌ Error: {e}"

# ------------------- File Selection Logic -------------------
def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
    if file_path:
        status_msg = highlight_formula_cells(file_path)
        status_var.set(status_msg)
        messagebox.showinfo("Process Completed", status_msg)
    else:
        status_var.set("⚠️ No file selected.")
        messagebox.showwarning("No Selection", "Please select a valid Excel file.")

# ------------------- GUI Layout -------------------
def launch_gui():
    global status_var

    root = tk.Tk()
    root.title("Sperene Excel Formula Highlighter")
    root.geometry("650x400")
    root.config(bg="#f0f8ff")

    # --- Header ---
    tk.Label(
        root, text="🔍 Formula Highlighter for Excel",
        font=("Helvetica", 18, "bold"),
        bg="#4682B4", fg="white", pady=12
    ).pack(fill=tk.X)

    # --- Description ---
    desc = (
        "👉 This tool allows you to:\n"
        "- Select an Excel file (.xlsx or .xls)\n"
        "- Automatically scan for formula cells\n"
        "- Highlight them in green\n"
        "- Instantly open the file after processing\n"
    )
    tk.Label(root, text=desc, justify="left", font=("Arial", 11),
             bg="#f0f8ff", fg="#333").pack(pady=20, padx=25, anchor="w")

    # --- Button ---
    btn_frame = tk.Frame(root, bg="#f0f8ff")
    btn_frame.pack()
    ttk.Button(btn_frame, text="📂 Select Excel File", command=select_excel_file).pack(pady=15, ipadx=15, ipady=5)

    # --- Status Bar ---
    status_var = tk.StringVar()
    status_bar = tk.Label(root, textvariable=status_var, anchor="w", bg="#d3eaff", font=("Arial", 11))
    status_bar.pack(fill=tk.X, side=tk.BOTTOM, ipady=5)

    root.mainloop()

# ------------------- Run GUI -------------------
if __name__ == "__main__":
    launch_gui()
